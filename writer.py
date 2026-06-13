"""Results workbook — four sheets:
  Answers      — per-student answers vs the official key + marks calculation
  Marks        — the marks-per-student table (sorted)
  Per-question — item analysis across all students (% correct)
  Review       — absentees, no-match / low-confidence / fallback / mark anomalies
"""

from __future__ import annotations

import logging

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config

PCT_FMT = '0.0"%"'
_SCORE_SCALE = ColorScaleRule(
    start_type="num", start_value=0, start_color="F8696B",      # red
    mid_type="num", mid_value=50, mid_color="FFEB84",           # yellow
    end_type="num", end_value=100, end_color="63BE7B",          # green
)
_THIN = Side(style="thin", color="D9D9D9")
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
# Homeroom→EMP equivalents (confirmed from sheets where students wrote BOTH
# notations) — used only to group the Summary's per-class means under EMP1–3.
_EMP_OF = {"A2801": "EMP1", "A2803": "EMP2", "A2804": "EMP3"}
_EMP_CLASSES = ("EMP1", "EMP2", "EMP3")


def _pin_title_top(chart):
    """Pin the chart title flush to the top edge (y=0), overlaying the plot."""
    chart.title.overlay = True
    chart.title.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.3, y=0.0))

log = logging.getLogger("xuekao")

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
GREY = PatternFill("solid", fgColor="E7E6E6")
HEADF = Font(color="FFFFFF", bold=True)
HEAD = PatternFill("solid", fgColor="305496")
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")

_TAG_FILL = {"correct": GREEN, "wrong": RED, "over": RED, "under": RED, "blank": YELLOW}


def _name_disp(rec: dict) -> str:
    return rec["name"]["matched_name"] or "?"


def _class_disp(rec: dict) -> str:
    # Never surface internal sentinels in the Class column — a blank field,
    # unmatched, or unreadable class all show as empty (the label isn't important).
    norm = rec["cls"]["class_norm"]
    return "" if norm in ("NONE", "NOMATCH", "UNREADABLE") else (norm or "")


def _serial(name: str, roster: list[str]) -> str:
    try:
        return str(roster.index(name) + 1)
    except ValueError:
        return ""


def _header(ws, values, row=1):
    for j, v in enumerate(values, 1):
        c = ws.cell(row=row, column=j, value=v)
        c.fill = HEAD
        c.font = HEADF
        c.alignment = CENTER


def _answers_sheet(ws, records, key, roster):
    ws.title = "Answers"
    qs = [it["q"] for it in key["questions"]]
    total = key["total_points"]
    head = (["序号", "Sheet", "Name", "Class"] + [f"Q{q}" for q in qs]
            + ["#Correct", f"Score/{total:g}", "%"])
    _header(ws, head)
    # official-answer reference row
    ws.cell(row=2, column=3, value="official answers →").font = BOLD
    for j, it in enumerate(key["questions"], 5):
        c = ws.cell(row=2, column=j, value="".join(it["correct"]))
        c.font = BOLD
        c.fill = GREY
        c.alignment = CENTER
    ws.cell(row=2, column=5 + len(qs) + 1, value=f"pts {total:g}").font = BOLD

    r = 3
    for rec in sorted(records, key=lambda x: x["page"]):
        score = rec["score"]
        name = _name_disp(rec)
        ws.cell(row=r, column=1, value=_serial(name, roster))
        ws.cell(row=r, column=2, value=rec["page"])
        ws.cell(row=r, column=3, value=name)
        ws.cell(row=r, column=4, value=_class_disp(rec))
        for j, q in enumerate(qs, 5):
            marked, tag, _pts = score.per_q.get(q, (frozenset(), "blank", 0.0))
            c = ws.cell(row=r, column=j, value="".join(sorted(marked)) or "·")
            c.alignment = CENTER
            c.fill = _TAG_FILL.get(tag, YELLOW)
        ws.cell(row=r, column=5 + len(qs), value=score.n_correct).alignment = CENTER
        ws.cell(row=r, column=5 + len(qs) + 1, value=score.total).alignment = CENTER
        pct = round(score.total / total * 100, 1) if total else 0.0
        pc = ws.cell(row=r, column=5 + len(qs) + 2, value=pct)
        pc.alignment = CENTER
        pc.number_format = PCT_FMT
        r += 1

    ws.freeze_panes = "E3"
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 9
    for j in range(5, 5 + len(qs)):  # narrow A/B/C/D answer columns
        ws.column_dimensions[get_column_letter(j)].width = 4.5


def _marks_sheet(ws, records, roster, max_points):
    # "Sheet" = the page in first_page.pdf, so a student's paper is one lookup away.
    _header(ws, ["Rank", "序号", "Sheet", "Name", "Class", "Score", "%", "#Correct", "Flags"])
    rows = []
    for rec in records:
        rows.append((rec["score"].total, rec["score"].n_correct, rec["page"],
                     _name_disp(rec), _class_disp(rec), rec))
    rows.sort(key=lambda t: (-t[0], -t[1], t[3]))
    for i, (total, ncorr, page, name, cls, rec) in enumerate(rows, 1):
        flags = "; ".join(_review_flags(rec))
        pct = round(total / max_points * 100, 1) if max_points else 0.0
        ws.append([i, _serial(name, roster), page, name, cls, total, pct, ncorr, flags])
    last = ws.max_row
    for row in range(2, last + 1):
        ws.cell(row=row, column=7).number_format = PCT_FMT
    if last >= 2:
        ws.conditional_formatting.add(f"G2:G{last}", _SCORE_SCALE)  # color scale on %
        ws.auto_filter.ref = f"A1:I{last}"                          # sortable/filterable
    ws.freeze_panes = "A2"
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["I"].width = 50


def _perq_sheet(ws, stats, key):
    _header(ws, ["Q", "Correct", "Points", "#Answered", "#Correct", "% Correct", "#Blank", "#Wrong"])
    for s in stats:
        ws.append([s["q"], s["correct"], s["points"], s["answered"],
                   s["n_correct"], s["pct"], s["blank"], s["wrong"]])
        row = ws.max_row
        for col in range(1, 9):
            cc = ws.cell(row=row, column=col)
            cc.alignment = CENTER
            cc.border = _BOX
        ws.cell(row=row, column=6).number_format = PCT_FMT
    data_last = ws.max_row
    if data_last >= 2:
        ws.conditional_formatting.add(f"F2:F{data_last}", _SCORE_SCALE)  # smooth gradient
        ws.auto_filter.ref = f"A1:H{data_last}"
        chart = BarChart()
        chart.type = "col"
        chart.title = "% correct by question"
        chart.legend = None
        chart.height, chart.width = 9, 24
        chart.y_axis.scaling.min, chart.y_axis.scaling.max = 0, 100
        chart.x_axis.delete = chart.y_axis.delete = False  # keep tick labels + value numbers
        _pin_title_top(chart)
        chart.add_data(Reference(ws, min_col=6, min_row=1, max_row=data_last), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=data_last))
        ws.add_chart(chart, "J2")
    ws.append([])
    ws.append(["TOTAL", "", key["total_points"], "", "", "", "", ""])
    ws.cell(row=ws.max_row, column=1).font = BOLD
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 8


def _review_flags(rec: dict) -> list[str]:
    flags = []
    nm, cl, an = rec["name"], rec["cls"], rec["ans"]
    name_flagged = nm["matched_name"] in ("NONAME", "UNREADABLE", "NOMATCH")
    if name_flagged:
        flags.append(f"name={nm['matched_name']}(raw={nm.get('raw_name', '')!r})")
    elif nm["confidence"] <= config.LOW_CONFIDENCE:
        flags.append(f"name low-conf={nm['confidence']}")
        name_flagged = True
    # Only surface the AI's name note when the name still needs attention — not
    # for confident matches or names we've fixed by hand (override clears it).
    if name_flagged and nm.get("problem"):
        flags.append(f"name note: {nm['problem']}")
    # Class is flagged ONLY when the AI struggled to READ it (low confidence).
    # A confidently-read class — including a confidently-blank field or a clear
    # but unmappable scrawl — needs no remark; the class label isn't important.
    if cl["confidence"] <= config.LOW_CONFIDENCE:
        flags.append(f"class low-conf={cl['confidence']}(raw={cl.get('class_raw', '')!r})")
    # Answer note only when the read itself is uncertain — routine under/over
    # marking is genuine (correct 0 per 少选/多选不得分) and shown tersely below.
    ans_low = an["confidence"] <= config.LOW_CONFIDENCE
    if ans_low:
        flags.append(f"answers low-conf={an['confidence']}")
        if an.get("problem"):
            flags.append(f"answers note: {an['problem']}")
    if rec.get("geom") and rec["geom"].get("frame_source") not in (None, "marks"):
        flags.append(f"anchor fallback={rec['geom']['frame_source']}")
    # Under-/over-filling is a STUDENT issue (correctly scored 0 per 少选/多选不得分),
    # NOT an AI read problem, so it is not flagged. Only read-coverage gaps
    # (missing/extra question numbers) are surfaced — those mean the AI didn't
    # read every question.
    score = rec.get("score")
    if score:
        flags.extend(score.problems)
    return flags


def _review_sheet(ws, records, key, roster, low_conf, full_run=True):
    _header(ws, ["Type", "序号", "Sheet", "Name", "Class", "N.conf", "C.conf", "A.conf", "Detail"])
    # key problems first
    if not key.get("valid", False):
        ws.append(["KEY", "", "", "", "", "", "", key.get("confidence", ""),
                   "; ".join(key.get("problems", [])) or "invalid key"])
    elif key.get("problem"):
        ws.append(["KEY", "", "", "", "", "", "", key.get("confidence", ""),
                   f"key note: {key['problem']}"])
    # absentees: roster names with no matched sheet — only meaningful on a full run
    matched = {_name_disp(r) for r in records
               if r["name"]["matched_name"] not in ("NONAME", "UNREADABLE", "NOMATCH")}
    if full_run:
        for nm in roster:
            if nm not in matched:
                ws.append(["ABSENT", _serial(nm, roster), "", nm, "", "", "", "",
                           "no sheet matched this roster name"])
    else:
        ws.append(["INFO", "", "", "", "", "", "", "",
                   f"partial run ({len(records)} sheets) — absentee list suppressed "
                   "(run the full folder to list absentees)"])
    # duplicate matches (same roster name on >1 sheet)
    from collections import Counter
    cnt = Counter(_name_disp(r) for r in records
                  if r["name"]["matched_name"] not in ("NONAME", "UNREADABLE", "NOMATCH"))
    dups = {n for n, c in cnt.items() if c > 1}
    # per-record flags
    for rec in sorted(records, key=lambda x: x["page"]):
        flags = _review_flags(rec)
        if _name_disp(rec) in dups:
            flags.insert(0, "DUPLICATE name match")
        if flags:
            ws.append(["FLAG", _serial(_name_disp(rec), roster), rec["page"],
                       _name_disp(rec), _class_disp(rec), rec["name"]["confidence"],
                       rec["cls"]["confidence"], rec["ans"]["confidence"], "; ".join(flags)])
    if ws.max_row >= 2:
        ws.auto_filter.ref = f"A1:I{ws.max_row}"
    ws.freeze_panes = "A2"
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["I"].width = 70


def _summary_sheet(ws, records, key):
    import statistics as st
    from collections import defaultdict
    total = key["total_points"]
    scores = [r["score"].total for r in records]
    pcts = [s / total * 100 for s in scores] if total else []
    n = len(records)

    ws["A1"] = f"{key.get('subject', '')} — {key['total_questions']} MCQ · {total:g} points".lstrip(" —·")
    ws["A1"].font = Font(bold=True, size=14, color="305496")

    # --- key stats block (A3:B…) ---
    items = [("Students graded", n, None)]
    if scores:
        items += [("Mean score", round(st.mean(scores), 1), None),
                  ("Mean %", round(st.mean(pcts), 1), PCT_FMT),
                  ("Median score", round(st.median(scores), 1), None),
                  ("Highest score", max(scores), None),
                  ("Lowest score", min(scores), None)]
    r = 3
    for label, val, fmt in items:
        lc = ws.cell(r, 1, label); lc.font = BOLD; lc.fill = GREY; lc.border = _BOX
        vc = ws.cell(r, 2, val); vc.alignment = CENTER; vc.border = _BOX
        if fmt:
            vc.number_format = fmt
        r += 1

    # --- score distribution (table A-B) + horizontal bar chart ---
    dist_hdr = r + 1
    ws.cell(dist_hdr, 1, "Score distribution").font = Font(bold=True, size=12)
    d0 = dist_hdr + 1
    bands = [("90–100%", 90, 101), ("80–89%", 80, 90), ("70–79%", 70, 80),
             ("60–69%", 60, 70), ("50–59%", 50, 60), ("below 50%", 0, 50)]
    for i, (label, lo, hi) in enumerate(bands):
        cnt = sum(1 for p in pcts if lo <= p < hi)
        ws.cell(d0 + i, 1, label).border = _BOX
        cc = ws.cell(d0 + i, 2, cnt); cc.alignment = CENTER; cc.border = _BOX
    d_last = d0 + len(bands) - 1
    ch1 = BarChart()
    ch1.type, ch1.legend, ch1.title = "bar", None, "Students by score band"
    ch1.height, ch1.width = 7, 13
    ch1.x_axis.delete = ch1.y_axis.delete = False
    _pin_title_top(ch1)
    ch1.add_data(Reference(ws, min_col=2, min_row=d0, max_row=d_last))
    ch1.set_categories(Reference(ws, min_col=1, min_row=d0, max_row=d_last))
    ws.add_chart(ch1, "D2")

    # --- by class (table A-D) + mean-% chart + color scale ---
    bc_hdr = d_last + 2
    for j, h in enumerate(("Class", "n", "Mean score", "Mean %"), 1):
        c = ws.cell(bc_hdr, j, h); c.font = HEADF; c.fill = HEAD; c.alignment = CENTER
    by = defaultdict(list)
    for rec in records:
        emp = _EMP_OF.get(_class_disp(rec), _class_disp(rec))  # merge A28xx→EMP
        if emp in _EMP_CLASSES:                                # show only EMP classes
            by[emp].append(rec["score"].total)
    classes = [c for c in _EMP_CLASSES if c in by]
    c0 = bc_hdr + 1
    for i, cls in enumerate(classes):
        ss = by[cls]
        mean = round(st.mean(ss), 1)
        ws.cell(c0 + i, 1, cls).border = _BOX
        for col, val in ((2, len(ss)), (3, mean)):
            cc = ws.cell(c0 + i, col, val); cc.alignment = CENTER; cc.border = _BOX
        pc = ws.cell(c0 + i, 4, round(mean / total * 100, 1) if total else 0)
        pc.number_format = PCT_FMT; pc.alignment = CENTER; pc.border = _BOX
    c_last = c0 + len(classes) - 1
    if c_last >= c0:
        ws.conditional_formatting.add(f"D{c0}:D{c_last}", _SCORE_SCALE)
        ch2 = BarChart()
        ch2.type, ch2.legend, ch2.title = "bar", None, "Mean % by class"
        ch2.height, ch2.width = 8, 13
        ch2.x_axis.delete = ch2.y_axis.delete = False
        _pin_title_top(ch2)
        ch2.add_data(Reference(ws, min_col=4, min_row=c0, max_row=c_last))
        ch2.set_categories(Reference(ws, min_col=1, min_row=c0, max_row=c_last))
        ws.add_chart(ch2, "F18")

    for colw, w in (("A", 18), ("B", 11), ("C", 12), ("D", 10)):
        ws.column_dimensions[colw].width = w


def write_workbook(records, key, stats, roster, out_path, low_conf=None, full_run=True):
    if low_conf is None:
        low_conf = config.LOW_CONFIDENCE
    wb = openpyxl.Workbook()
    _answers_sheet(wb.active, records, key, roster)
    _marks_sheet(wb.create_sheet("Marks"), records, roster, key["total_points"])
    _perq_sheet(wb.create_sheet("Per-question"), stats, key)
    _review_sheet(wb.create_sheet("Review"), records, key, roster, low_conf, full_run)
    _summary_sheet(wb.create_sheet("Summary", 0), records, key)  # first tab = overview
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    log.info("wrote %s (%d students)", out_path.name, len(records))
